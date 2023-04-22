from openpyxl import load_workbook
from openpyxl import Workbook

def updateValues(height=0.0, outer=0.0, inner=0.0):

    wb = load_workbook('C:\\autofrp\\123.xlsx')
    ws = wb['Sheet1']
    ws['C3'] = height
    ws['F3'] = outer
    ws['G3'] = inner
    wb.save('C:\\autofrp\\123.xlsx')