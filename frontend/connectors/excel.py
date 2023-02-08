from openpyxl import load_workbook
from openpyxl import Workbook

def updateValues(filletradius, height, top_lid, bottom_lid, outer_radius, inner_radius):
    wb = load_workbook('C:\\autofrp\\table.xlsx')
    ws = wb['Sheet1']
    ws['B3'] = filletradius
    ws['C3'] = height
    ws['D3'] = top_lid
    ws['E3'] = bottom_lid
    ws['F3'] = outer_radius
    ws['G3'] = inner_radius
    wb.save('C:\\autofrp\\table.xlsx')